from openpyxl import Workbook
from typing import List
import os
from app.models.stock_data_model import StockDataModel

class ExcelExportService:
    def export_to_excel(self, data_list: List[StockDataModel], file_path: str):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Stock Data"

        # Crear el encabezado de la tabla
        columns = [
            "Nemotécnico", "Último Precio", "Variación Porcentual",
            "Volúmenes", "Cantidad", "Variación Absoluta",
            "Precio Apertura", "Precio Máximo", "Precio Mínimo",
            "Precio Promedio", "Nombre Emisor"
        ]
        
        # Escribir encabezados
        for col, header in enumerate(columns, 1):
            sheet.cell(row=1, column=col, value=header)

        # Escribir datos
        for row, stock_data in enumerate(data_list, 2):
            data = stock_data.dict()
            for col, key in enumerate(data.keys(), 1):
                sheet.cell(row=row, column=col, value=data[key])

        try:
            workbook.save(file_path)
            
            # Abrir el archivo en Windows
            if os.name == 'nt':
                os.startfile(os.path.abspath(file_path))
                
        except Exception as e:
            raise RuntimeError(f"Error al escribir archivo Excel: {str(e)}")